#!/usr/bin/env python3
"""LQA Inspector - Severity Rules Parser.

Deterministic XLSX parser for canonical severity rule workbooks.
Uses openpyxl to parse BUG info sheets with merged-cell carry-forward.
"""

import json
import os
import sys
from typing import Optional

# Ensure UTF-8 stdout on Windows (cp1252 cannot encode CJK/emoji/Indonesian chars)
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')


def parse_workbook(workbook_path: str) -> dict:
    import openpyxl

    if not os.path.isfile(workbook_path):
        return {"success": False, "error": f"Workbook not found: {workbook_path}", "rules": [], "warnings": []}

    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
    except Exception as e:
        return {"success": False, "error": f"Failed to open workbook: {e}", "rules": [], "warnings": []}

    bug_sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == "bug info":
            bug_sheet_name = name
            break

    if bug_sheet_name is None:
        return {"success": False, "error": f"No BUG info sheet found. Available: {wb.sheetnames}", "rules": [], "warnings": []}

    ws = wb[bug_sheet_name]
    filename = os.path.basename(workbook_path)
    warnings: list[str] = []
    rules: list[dict] = []
    merged_ranges = list(ws.merged_cells.ranges)

    def get_merged_value(row: int, col: int) -> Optional[str]:
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            return str(cell.value)
        for mr in merged_ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
                top_left = ws.cell(row=mr.min_row, column=mr.min_col)
                if top_left.value is not None:
                    return str(top_left.value)
        return None

    header_row = 2
    for r in range(1, min(10, ws.max_row + 1)):
        val = ws.cell(row=r, column=2).value
        if val and "sub" in str(val).lower():
            header_row = r
            break

    data_start = header_row + 1
    current_category = None
    current_subcategory = None

    for row_idx in range(data_start, ws.max_row + 1):
        cat_val = get_merged_value(row_idx, 1)
        if cat_val is not None:
            cat_val = cat_val.strip()
            if len(cat_val) <= 200:
                current_category = cat_val
            else:
                warnings.append(f"Row {row_idx} Col A: skipped long prose ({len(cat_val)} chars)")

        sub_val = get_merged_value(row_idx, 2)
        if sub_val is not None:
            current_subcategory = sub_val.strip()

        severity_cell = ws.cell(row=row_idx, column=3)
        severity_val = severity_cell.value
        if severity_val is None or str(severity_val).strip() == "":
            continue
        severity_val = str(severity_val).strip()

        desc_cell = ws.cell(row=row_idx, column=4)
        description = str(desc_cell.value).strip() if desc_cell.value is not None else ""

        if current_subcategory is None:
            warnings.append(f"Row {row_idx}: no sub-category, skipping")
            continue

        parts = severity_val.split(None, 1)
        severity_code = ""
        severity_name = ""
        if parts and parts[0].startswith("P") and parts[0][1:].isdigit():
            severity_code = parts[0]
            severity_name = parts[1] if len(parts) > 1 else ""

        rules.append({
            "source_row": row_idx,
            "main_category_raw": current_category or "",
            "subcategory_raw": current_subcategory,
            "canonical_bug_type": current_subcategory,
            "severity_exact": severity_val,
            "severity_code": severity_code,
            "severity_name": severity_name,
            "description_raw": description,
            "source_sheet": bug_sheet_name,
            "source_filename": filename,
        })

    wb.close()
    return {
        "success": True,
        "error": None,
        "rules": rules,
        "warnings": warnings,
        "metadata": {"filename": filename, "sheet_name": bug_sheet_name, "total_rules": len(rules)},
    }


def main():
    if len(sys.argv) < 2:
        print(json.dumps({"success": False, "error": "Usage: severity_rules_parser.py <workbook_path>"}))
        sys.exit(1)
    result = parse_workbook(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False))


if __name__ == "__main__":
    main()

